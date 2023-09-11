import streamlit as st
import numpy as np
import pandas as pd
import openpyxl
import plotly.graph_objects as go
from plotly.subplots import make_subplots

st.set_page_config(layout='wide')


@st.cache_data
def load_data(meas_file, data_file):
    meas_wb = openpyxl.load_workbook(meas_file)
    data_wb = openpyxl.load_workbook(data_file)
    wells = list(set(meas_wb.sheetnames) & set(data_wb.sheetnames))
    measurements = {}
    computations = {}
    for well in wells:
        measurements[well] = pd.read_excel(meas_file, sheet_name=well, index_col=0)
        computations[well] = pd.read_excel(data_file, sheet_name=well)
    #st.write("В замерах и результатах :red[нет совпадающих скважин! Загрузите другие файлы.]")
    return wells, measurements, computations

st.sidebar.header("Загрузите Excel-файлы с данными")
meas_file = st.sidebar.file_uploader("Выберите файл с замерами", type=['xlsx'])
data_file = st.sidebar.file_uploader("Выберите файл с расчетами", type=['xlsx'])
if (meas_file is not None) and (data_file is not None):
    wells, measurements, computations = load_data(meas_file, data_file)
    if wells is not None:
        st.sidebar.markdown("----")
        st.sidebar.markdown("<h3>Выберите скважину</h3>", unsafe_allow_html=True)
        selected_well = st.sidebar.selectbox('', wells, key='dyn-well-selector')
        well_measurments = measurements[selected_well]
        well_computed = computations[selected_well]
        min_depth = well_measurments['Abs'].min()
        cols = well_measurments.columns[1:]
        indices = np.where(well_computed['Abs'] >= min_depth)

        st.subheader(f"Скважина {selected_well}")
        tab1, tab2 = st.tabs(["🟥 Динамика", "📈 Одномерные графики"])

        with tab1:
            contours = go.contour.Contours(
                coloring='heatmap',
                showlabels=True,
                start=-5,
                end=7,
                size=0.5
            )

            computation = go.Contour(
                z = well_computed[cols[1:]].values[indices],
                y = well_computed['Abs'].loc[indices],
                x = cols.values,
                colorscale='Jet',
                contours=contours)

            measurements = go.Contour(
                z = well_measurments[cols[1:]].values,
                y = well_measurments['Abs'],
                x = cols.values,
                colorscale='Jet',
                contours=contours)

            fig1 = make_subplots(
                rows=2, cols=1,
                shared_xaxes=True,
                #    shared_yaxes=True,
                    # vertical_spacing=0.03,
                subplot_titles=('Рассчитанные значения', 'Замеры')
            )

            fig1.add_trace(
                computation,
                row=1, col=1
            )

            fig1.add_trace(
                measurements,
                row=2, col=1)

            fig1.update_xaxes(title_text='Даты измерений')
            fig1.update_yaxes(title_text='Абсолютная отметка, м')
            fig1.update_layout(width=1000, height=800)
            st.plotly_chart(fig1, theme=None)

        with tab2:
            date = st.select_slider(
                "Дата",
                options=cols
            )

            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                x = well_computed[date],
                y = well_computed['Abs'].loc[indices],
                name = 'Расчет'
            ))

            fig2.add_trace(go.Scatter(
                x = well_measurments[date],
                y = well_measurments['Abs'],
                name = 'Замеры',
                mode='lines+markers'
            ))

            fig2.update_xaxes(title_text='Температура', range=[-10, 10])
            fig2.update_yaxes(title_text='Абсолютная отметка, м')
            fig2.update_layout(width=1000, height=800)
            st.plotly_chart(fig2, theme=None)
